from odoo import _, models, fields, api
from odoo.exceptions import UserError, ValidationError
import base64
import io
import csv
import xlsxwriter
from openpyxl import load_workbook

class ImportContactWizard(models.TransientModel):
    _name = 'import.contact.wizard'
    _description = 'Import Contact Wizard'

    file = fields.Binary(string="Archivo", required=False)
 
    def action_test(self):
        file = self.file
        if not file:
            raise UserError(_("No se ha subido ningún archivo. Por favor, sube un archivo CSV o XLSX."))

        file_data = base64.b64decode(file)
        file_format = self._get_file_format(file_data)

        if file_format == 'csv':
            lines = file_data.decode('utf-8').splitlines()
            reader = csv.reader(lines)
            next(reader)  # Obtener encabezado
            data = list(reader)
        elif file_format == 'xlsx':
            workbook = load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
        else:
            raise UserError(_("Formato de archivo no válido. Solo se permiten archivos CSV o XLSX."))
        errors = self.validate_data(data)
    

        if errors:
            raise UserError(_("Errores encontrados en los registros:\n\n" + "\n".join(errors) ))
        else:
          
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Validación exitosa',
                    'message': '¡Todo parece correcto en el archivo de productos!',
                    'type': 'success',
                    'sticky': False,
                }
            }

    def _get_file_format(self, file_data):
        """Detecta el tipo de archivo según su contenido"""
        try:
            if file_data[:4] == b'PK\x03\x04':  # Este es el encabezado de archivos XLSX
                return 'xlsx'
            elif file_data[:3] == b'\xef\xbb\xbf':  # BOM de archivos CSV (UTF-8 con BOM)
                return 'csv'
        except Exception:
            pass
        return ''  

    def validate_data(self, data):
        errors = []
        for i, row in enumerate(data, start=2):
            if len(row) < 6:
                errors.append(f"Fila {i}: Datos incompletos.")
                continue

            otra_direccion, _, _, ciudad, _, _ = row[:6]

            if otra_direccion not in ['Contacto', 'Dirección de factura', 'Dirección de entrega', 'Otra dirección']:
                errors.append(f"Fila {i}: El campo 'Otra Dirección' tiene un valor no válido.")
            if not ciudad:
                errors.append(f"Fila {i}: El campo 'Ciudad' no debe estar vacío.")
        return errors
        
    def action_export_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        # Hoja principal con encabezados
        worksheet = workbook.add_worksheet('Plantilla Contactos')

        header_format = workbook.add_format({
            'bold': True, 'text_wrap': True, 'valign': 'vcenter',
            'align': 'center', 'border': 1, 'bg_color': '#D9E1F2'
        })

        headers = ['Nombre', 'Es Empresa', 'Es Cliente', 'Es Proveedor',
                   'Otra Dirección', 'Calle', 'Calle 2', 'Ciudad', 'Estado', 'País']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        worksheet.set_column(0, len(headers) - 1, 25)

        # observaciones
        worksheet2 = workbook.add_worksheet('Observaciones')
        header_observations_format = workbook.add_format({
            'bold': True, 'text_wrap': True, 'valign': 'vcenter',
            'align': 'center', 'border': 1, 'bg_color': '#4CAF50', 'color': 'white'
        })

        observations_format = workbook.add_format({
            'bold': False, 'text_wrap': True, 'valign': 'top',
            'align': 'left', 'border': 1, 'bg_color': '#F9F9F9'
        })

        worksheet2.write('A1', 'Observaciones', header_observations_format)

        observaciones = [
            'Nombre: Obligatorio. Nombre del contacto o empresa.',
            'Es Empresa: Escriba True si es una empresa, False si es una persona.',
            'Es Cliente: True si el contacto será cliente.',
            'Es Proveedor: True si el contacto será proveedor.',
            'Otra Dirección: Use Contacto, Dirección de factura, Dirección de entrega u Otra dirección. Solo aplica si Es Empresa = False.',
            'Calle, Calle 2, Ciudad, Estado, País: Campos de dirección del contacto.'
        ]

        for row, obs in enumerate(observaciones, start=1):
            worksheet2.write(row, 0, obs, observations_format)

        worksheet2.set_column('A:A', 80)

        workbook.close()
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_contactos.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'import.contact.wizard',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }   

    def action_import(self):
        if not self.file:
            raise UserError(_("Debe subir un archivo."))

        file_data = base64.b64decode(self.file)
        file_format = self._get_file_format(file_data)
        data = self._parse_file(file_data, file_format)

        errores = []
        duplicates = []
        contactos_creados = 0
        
        for i, row in enumerate(data, start=2):
            try:
                nombre, es_empresa_str, es_cliente_str, es_proveedor_str, otra_direccion, calle, calle_2, ciudad, estado, pais = row

                # Convertir strings a booleanos
                is_company = str(es_empresa_str).strip().lower() == 'true'
                is_customer = str(es_cliente_str).strip().lower() == 'true'
                is_supplier = str(es_proveedor_str).strip().lower() == 'true'

                # Verificar si ya existe un contacto igual
                existing_contact = self.env['res.partner'].search([
                    ('name', '=', nombre),
                    ('is_company', '=', is_company)
                ], limit=1)

                if existing_contact:
                    duplicates.append(f"row {i}: {nombre}")
                    continue 

                country = self.env['res.country'].search([('name', 'ilike', pais)], limit=1)
                state = self.env['res.country.state'].search([('name', 'ilike', estado)], limit=1)
                
                vals = {
                    'name': nombre,
                    'is_company': is_company,
                    'type': self._map_direccion_type(otra_direccion),
                    'street': calle,
                    'street2': calle_2,
                    'city': ciudad,
                    'state_id': state.id if state else False,
                    'country_id': country.id if country else False,
                    'customer_rank': 1 if is_customer else 0,
                    'supplier_rank': 1 if is_supplier else 0,
                }
                

                self.env['res.partner'].create(vals)
                contactos_creados += 1

            except Exception as e:
                errores += f"Fila {i}: Error al crear contacto: {str(e)}\n"

        if errores:
            raise UserError(_("Errores durante la importación:\n\n" + errores))
        
        if duplicates:
            mensaje = "\n".join(duplicates)
            raise UserError(_("Los siguientes contactos ya existen y no fueron importados:\n\n" + mensaje))

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Importación completada',
                'message': f'Se importaron correctamente {contactos_creados} contactos.',
                'type': 'success',
                'sticky': False,
            }
        }

    def _map_direccion_type(self, tipo):
        mapping = {
            'Contacto': 'contact',
            'Dirección de factura': 'invoice',
            'Dirección de entrega': 'delivery',
            'Otra dirección': 'other',
        }
        return mapping.get(tipo, 'contact')

    
    def _parse_file(self, file_data, file_format):
        if file_format == 'csv':
            lines = file_data.decode('utf-8').splitlines()
            reader = csv.reader(lines)
            next(reader)
            return list(reader)
        elif file_format == 'xlsx':
            workbook = load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            return [row for row in sheet.iter_rows(min_row=2, values_only=True)]
        else:
            raise UserError(_("Formato de archivo no válido. Solo se permiten archivos CSV o XLSX."))


    