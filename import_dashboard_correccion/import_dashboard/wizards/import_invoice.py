import base64
import binascii
import csv
import datetime
import io
import tempfile
import openpyxl  # Para XLSX en lugar de xlrd
from odoo.exceptions import ValidationError
from odoo import fields, models


class ImportInvoice(models.TransientModel):
    """Model for import invoice"""
    _name = 'import.invoice'
    _description = 'Invoice Import'

    file_type = fields.Selection(
        selection=[('csv', 'CSV File'), ('xlsx', 'XLSX File')],
        string='Import File Type', default='csv',
        help="It helps to choose the file type"
    )
    file = fields.Binary(string="File", help="File")
    update_posted = fields.Boolean(
        string='Update Posted Record?',
        help='If enabled, the records in "Posted" state will be converted to draft'
             ' and values are updated. These records will then again be posted'
             ' if "Post Automatically" is activated'
    )
    auto_post = fields.Boolean(string='Post Automatically',
                               help="Post Automatically"
                               )
    journal = fields.Selection(
        selection=[('Bank', 'Bank'), ('Cash', 'Cash')],
        string='Journal', default='Bank', help='It helps to choose Journal type'
    )
    order_number = fields.Selection(
        selection=[('from_system', 'From System'), ('from_file', 'From File')],
        string='Number', default='from_file', help="Order number"
    )
    import_product_by = fields.Selection(
        selection=[('name', 'Name'), ('default_code', 'Internal Reference'),
                   ('barcode', 'Barcode')], required=True, default="name",
        string="Import invoice by", help="Product import"
    )
    type = fields.Selection(
        selection=[('out_invoice', 'Invoice'), ('in_invoice', 'Bill'),
                   ('out_refund', 'Credit Note'), ('in_refund', 'Refund')],
        string='Invoicing Type', required=True, help="Invoice type",
        default="out_invoice"
    )

    def action_import_invoice(self):
        """Creating Invoice record using uploaded xl/csv files"""
        account_move = self.env['account.move']
        res_partner = self.env['res.partner']
        res_users = self.env['res.users']
        account_account = self.env['account.account']
        uom_uom = self.env['uom.uom']
        account_tax = self.env['account.tax']
        product_product = self.env['product.product']
        product_attribute = self.env['product.attribute']
        product_attribute_value = self.env['product.attribute.value']
        product_template_attribute_value = self.env[
            'product.template.attribute.value']

        items = self.read_file()

        imported = 0
        confirmed = 0
        imported_invoices = []
        error_msg = ""
        partner_added_msg = ""
        warning_msg = ""

        for row, item in enumerate(items, start=1):
            vals = {}
            row_not_import_msg = "\n❌Row {rn} not imported.".format(rn=row)
            import_error_msg = ""
            missing_fields_msg = ""
            fields_msg = "\n\t🚫Missing required field(s):"
            partner_msg = "\n🆕New Partner(s) added:"
            vals['move_type'] = self.type

            # Partner Processing
            if item.get('Partner'):
                partner = res_partner.search([('name', '=', item['Partner'])])
                if not partner:
                    partner = res_partner.create({
                        'name': item['Partner']
                    })
                    vals['partner_id'] = partner.id
                    partner_added_msg += partner_msg + "\n\t\trow {rn}: \"{partner}\"".format(
                        rn=row, partner=item['Partner'])
                elif len(partner) > 1:
                    import_error_msg += row_not_import_msg + (
                        "\n\t\t⚠ Multiple Partners with name (%s) found!" % item['Partner'])
                else:
                    vals['partner_id'] = partner.id
            else:
                missing_fields_msg += (fields_msg + "\n\t\t❗ \"Partner\"")

            # Invoice Date Processing
            if item.get('Invoice Date'):
                try:
                    invoice_date = datetime.datetime.strptime(item['Invoice Date'], '%m/%d/%Y')
                    vals['invoice_date'] = invoice_date
                except:
                    import_error_msg += row_not_import_msg + "\n\t\t⚠ Invalid Invoice Date format."

            # Due Date Processing
            if item.get('Due Date'):
                try:
                    due_date = datetime.datetime.strptime(item['Due Date'], '%m/%d/%Y')
                    vals['invoice_date_due'] = due_date
                except:
                    import_error_msg += row_not_import_msg + "\n\t\t⚠ Invalid Due Date format."

            # Searching for existing invoices or creating new ones
            invoice = account_move.search([('name', '=', item.get('Number')),
                                           ('move_type', '=', vals['move_type'])])
            if invoice:
                if len(invoice) > 1:
                    error_msg += row_not_import_msg + "\n\t⚠ Multiple invoices with same Number (%s) found!" % item['Number']
                    continue
                if vals:
                    if self.update_posted and invoice.state == 'posted':
                        invoice.button_draft()
                        invoice.write(vals)
                    elif invoice.state == 'draft':
                        invoice.write(vals)
            else:
                if self.order_number == 'from_system':
                    invoice = account_move.create(vals)
                if self.order_number == 'from_file' and item.get('Number'):
                    vals['name'] = item['Number']
                    invoice = account_move.create(vals)
                else:
                    error_msg += row_not_import_msg + "\n\t⚠ Missing Invoice Number."
                    continue

            # Processing Invoice Lines
            line_vals = {}
            pro_vals = {}

            if item.get('Product'):
                product = product_product.search([('name', '=', item['Product'])])
                if not product:
                    pro_vals['name'] = item['Product']
                    product = product_product.create(pro_vals)

                line_vals['product_id'] = product.id
            else:
                error_msg += row_not_import_msg + "\n\t⚠ Product missing in file!"

            # Account code and UOM processing
            if item.get('Account Code'):
                account = account_account.search([('code', '=', int(item['Account Code']))])
                line_vals['account_id'] = account.id
            if item.get('Uom'):
                uom = uom_uom.search([('name', '=', item['Uom'])])
                if uom:
                    line_vals['product_uom_id'] = uom.id

            # Price and Quantity
            if item.get('Quantity'):
                line_vals['quantity'] = item['Quantity']
            if item.get('Price'):
                line_vals['price_unit'] = item['Price']

            # Adding Line to the Invoice
            if line_vals:
                invoice.write({'invoice_line_ids': [(0, 0, line_vals)]})
                imported += 1
                imported_invoices.append(invoice)

            if self.auto_post and imported_invoices:
                for inv in imported_invoices:
                    inv.action_post()
                    confirmed += 1

        if error_msg:
            error_msg = "\n\n🏮 WARNING 🏮" + error_msg
            error_message = self.env['import.message'].create({'message': error_msg})
            return {
                'name': 'Error!',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'import.message',
                'res_id': error_message.id,
                'target': 'new'
            }

        msg = f"Imported {imported} records.\nPosted {confirmed} records" + partner_added_msg + warning_msg
        message = self.env['import.message'].create({'message': msg})

        return {
            'effect': {
                'fadeout': 'slow',
                'message': msg,
                'type': 'rainbow_man',
            }
        }

    def action_test_import_invoice(self):
        """Test import file without importing data."""
        if not self.file:
            raise ValidationError("Por favor, sube un archivo válido antes de continuar.")

        # Validación del tipo de archivo
        if self.file_type not in ['csv', 'xlsx']:
            raise ValidationError("Invalid file type. Only CSV and XLSX are allowed.")

        try:
            if self.file_type == 'csv':
                # Verificar que el archivo CSV tenga contenido
                csv_data = base64.b64decode(self.file)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                data_file.seek(0)
                reader = csv.reader(data_file)
                rows = list(reader)
                if len(rows) <= 1:
                    raise ValidationError("The CSV file is empty or missing data.")
            elif self.file_type == 'xlsx':
                # Verificar que el archivo XLSX tenga contenido
                fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                fp.write(binascii.a2b_base64(self.file))
                fp.seek(0)
                workbook = openpyxl.load_workbook(fp.name)
                sheet = workbook.active
                if sheet.max_row <= 1:  # Si solo tiene una fila (cabecera), es un archivo vacío
                    raise ValidationError("The XLSX file is empty or missing data.")
        except Exception as e:
            raise ValidationError(f"Error reading the file: {str(e)}")

        # Si las validaciones son exitosas
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Validation Success',
                'message': 'The file was validated successfully.',
                'sticky': False,
            }
        }

    def read_file(self):
        """Method to read the file depending on its type"""
        if self.file_type == 'csv':
            return self.read_csv_file()
        elif self.file_type == 'xlsx':
            return self.read_xlsx_file()

    def read_csv_file(self):
        """Read a CSV file"""
        try:
            csv_data = base64.b64decode(self.file)
            data_file = io.StringIO(csv_data.decode("utf-8"))
            data_file.seek(0)
            csv_reader = csv.DictReader(data_file, delimiter=',')
            return list(csv_reader)
        except Exception as e:
            raise ValidationError(f"Archivo CSV no válido. Error: {e}")

    def read_xlsx_file(self):
        """Read an XLSX file"""
        try:
            fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            fp.write(binascii.a2b_base64(self.file))
            fp.seek(0)
            workbook = openpyxl.load_workbook(fp.name)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            return rows
        except Exception as e:
            raise ValidationError(f"Archivo XLSX no válido. Error: {e}")

    def action_generate_template(self):
        """Genera una plantilla Excel para importar facturas"""

        import base64
        from io import BytesIO
        import xlsxwriter

        # Definimos los encabezados requeridos
        field_labels = [
            "Socio",                   # Partner
            "Fecha de Factura",        # Invoice Date
            "Fecha de Vencimiento",    # Due Date
            "Número",                  # Number
            "Producto",                # Product
            "Código de Cuenta",        # Account Code
            "UoM",                     # Uom
            "Cantidad",                # Quantity
            "Precio"                   # Price
        ]

        # Crear archivo Excel en memoria
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Invoice Import Template")

        # Estilo para encabezado
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9'})
        for col, label in enumerate(field_labels):
            worksheet.write(0, col, label, header_format)

        workbook.close()
        output.seek(0)

        # Crear archivo adjunto en Odoo
        attachment = self.env['ir.attachment'].create({
            'name': 'invoice_import_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Retornar acción de descarga
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }