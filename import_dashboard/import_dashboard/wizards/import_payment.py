import base64
import binascii
import csv
import io
import tempfile
import openpyxl  # Usamos openpyxl en lugar de xlrd para archivos XLSX
from odoo import fields, models
from odoo.exceptions import ValidationError
from datetime import datetime

class ImportPayment(models.TransientModel):
    _name = 'import.payment'
    _description = 'Payment Import'

    file_type = fields.Selection(
        selection=[('csv', 'CSV File'), ('xlsx', 'XLSX File')],
        string='Select File Type', default='xlsx',
        help='It helps to choose the file type'
    )
    file_upload = fields.Binary(string='File Upload', attachment=False, help='It helps to upload files')
    payment_type = fields.Selection(
        selection=[('inbound', 'Receive Money'), ('outbound', 'Send Money')],
        string='Payment Type',
        help='Type of payment (Receive Money or Send Money)'
    )

    def action_import_payment(self):
        """ Method to import payments from .csv or .xlsx files. """
        # Validar si hay archivo subido
        if not self.file_upload:
            raise ValidationError("Por favor, sube un archivo válido antes de continuar.")
        
        if self.file_type == 'csv':
            try:
                # Intentamos leer como CSV
                csv_data = base64.b64decode(self.file_upload)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                data_file.seek(0)
                reader = csv.reader(data_file)
                rows = list(reader)
                if len(rows) <= 1:  # Si solo tiene una fila (cabecera), es un archivo vacío
                    raise ValidationError("El archivo CSV está vacío o falta contenido.")
            except Exception as e:
                raise ValidationError(f"Error al leer el archivo CSV: {str(e)}")
        
        elif self.file_type == 'xlsx':
            try:
                # Intentamos leer como XLSX usando openpyxl
                fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                fp.write(binascii.a2b_base64(self.file_upload))
                fp.seek(0)
                workbook = openpyxl.load_workbook(fp.name)
                sheet = workbook.active
                if sheet.max_row <= 1:  # Si solo tiene una fila (cabecera), es un archivo vacío
                    raise ValidationError("El archivo XLSX está vacío o falta contenido.")
                rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Comienza desde la segunda fila (excluye cabecera)
                    rows.append(row)
            except Exception as e:
                raise ValidationError(f"Error al leer el archivo XLSX: {str(e)}")
        
        else:
            raise ValidationError("Tipo de archivo no válido. Solo se permiten archivos CSV y XLSX.")

        # Validar campos requeridos
        error_msg = ""
        for row_index, row in enumerate(rows, start=2):  # Empieza desde la fila 2 (excluyendo la cabecera)
            if not row or len(row) < 6:
                error_msg += f"Fila {row_index} está vacía o incompleta.\n"
                continue

            if not row[0]:  # Verificar que Amount esté presente
                error_msg += f"El monto falta en la fila {row_index}\n"
            if not row[1]:  # Verificar que Date esté presente
                error_msg += f"La fecha falta en la fila {row_index}\n"
            if not row[2]:  # Verificar que Journal esté presente
                error_msg += f"El diario falta en la fila {row_index}\n"
            if not row[3]:  # Verificar que Customer/Vendor esté presente
                error_msg += f"Cliente/Proveedor falta en la fila {row_index}\n"
            if not row[4]:  # Verificar que Payment Type esté presente
                error_msg += f"Tipo de pago falta en la fila {row_index}\n"
            if not row[5]:  # Verificar que Number esté presente
                error_msg += f"El número de pago falta en la fila {row_index}\n"

        if error_msg:
            raise ValidationError(f"Error de validación:\n{error_msg}")

        # Validar tipo de pago
        for row_index, row in enumerate(rows, start=2):
            # Asegúrate de que se utilicen los valores correctos 'inbound' y 'outbound'
            if row[4] not in ['inbound', 'outbound']:  # 'inbound' y 'outbound' son los valores correctos
                error_msg += f"Tipo de pago inválido en la fila {row_index}\n"

        if error_msg:
            raise ValidationError(f"Error de validación:\n{error_msg}")

        # Si todas las validaciones pasaron, crear registros de pago
        for row_index, row in enumerate(rows, start=2):
            # Crear el pago con los datos importados
            amount = row[0]  # Monto
            
            # Manejar la fecha con flexibilidad
            date_str = str(row[1]).strip()
            try:
                if ' ' in date_str:
                    date_str = date_str.split()[0]
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                raise ValidationError(f"Formato de fecha inválido en la fila {row_index}. Use el formato YYYY-MM-DD")
            
            journal = row[2]  # Diario
            partner_name = row[3]  # Cliente/Proveedor
            payment_type = row[4]  # Tipo de pago (Receive o Send)
            payment_number = row[5]  # Número de pago

            # Crear el partner si no existe
            partner = self.env['res.partner'].search([('name', '=', partner_name)], limit=1)
            if not partner:
                partner = self.env['res.partner'].create({'name': partner_name})

            # Crear el registro de pago con el campo correcto 'date'
            self.env['account.payment'].create({
                'amount': amount,
                'date': date,  # Usamos el campo 'date' correcto
                'journal_id': self.env['account.journal'].search([('name', '=', journal)], limit=1).id,
                'partner_id': partner.id,
                'payment_type': payment_type,
                'name': payment_number,
            })

        # Notificación de éxito
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Importación exitosa',
                'message': 'Los pagos se importaron correctamente.',
                'sticky': False,
            }
        }

    def action_test_import_payment(self):
        """Test the validity of the file before import."""
        if not self.file_upload:
            raise ValidationError("Por favor, sube un archivo válido antes de continuar.")

        # Validar tipo de archivo
        if self.file_type == 'csv':
            try:
                csv_data = base64.b64decode(self.file_upload)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                data_file.seek(0)
                reader = csv.reader(data_file)
                rows = list(reader)
                if len(rows) <= 1:  # Si solo tiene una fila (cabecera), es un archivo vacío
                    raise ValidationError("El archivo CSV está vacío o falta contenido.")
            except Exception as e:
                raise ValidationError(f"Error al leer el archivo CSV: {str(e)}")

        elif self.file_type == 'xlsx':
            try:
                fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                fp.write(binascii.a2b_base64(self.file_upload))
                fp.seek(0)
                workbook = openpyxl.load_workbook(fp.name)
                sheet = workbook.active
                if sheet.max_row <= 1:  # Si solo tiene una fila (cabecera), es un archivo vacío
                    raise ValidationError("El archivo XLSX está vacío o falta contenido.")
                rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Comienza desde la segunda fila (excluye cabecera)
                    rows.append(row)
            except Exception as e:
                raise ValidationError(f"Error al leer el archivo XLSX: {str(e)}")
        
        else:
            raise ValidationError("Tipo de archivo no válido. Solo se permiten archivos CSV y XLSX.")

        # Validar campos requeridos
        error_msg = ""
        for row_index, row in enumerate(rows, start=2):
            if not row or len(row) < 6:
                error_msg += f"Fila {row_index} está vacía o incompleta.\n"
                continue

            if not row[0]:  # Verificar que Amount esté presente
                error_msg += f"El monto falta en la fila {row_index}\n"
            if not row[1]:  # Verificar que Date esté presente
                error_msg += f"La fecha falta en la fila {row_index}\n"
            if not row[2]:  # Verificar que Journal esté presente
                error_msg += f"El diario falta en la fila {row_index}\n"
            if not row[3]:  # Verificar que Customer/Vendor esté presente
                error_msg += f"Cliente/Proveedor falta en la fila {row_index}\n"
            if not row[4]:  # Verificar que Payment Type esté presente
                error_msg += f"Tipo de pago falta en la fila {row_index}\n"
            if not row[5]:  # Verificar que Number esté presente
                error_msg += f"El número de pago falta en la fila {row_index}\n"

        if error_msg:
            raise ValidationError(f"Error de validación:\n{error_msg}")

        # Validar tipo de pago
        for row_index, row in enumerate(rows, start=2):
            # Asegúrate de que se utilicen los valores correctos 'inbound' y 'outbound'
            if row[4] not in ['inbound', 'outbound']:  # 'inbound' y 'outbound' son los valores correctos
                error_msg += f"Tipo de pago inválido en la fila {row_index}\n"

        if error_msg:
            raise ValidationError(f"Error de validación:\n{error_msg}")

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Validación exitosa',
                'message': 'El archivo fue validado exitosamente.',
                'sticky': False,
            }
        }

    def action_generate_template(self):
        """Genera una plantilla Excel para importar pagos"""

        import base64
        from io import BytesIO
        import xlsxwriter

        # Definimos los encabezados requeridos en el archivo de importación
        field_labels = [
            "Monto",                  # Amount
            "Fecha",                    # Date
            "Diario",                 # Journal
            "Cliente/Proveedor",         # Customer/Vendor
            "Tipo de Pago",            # Payment Type
            "Número"                   # Number
        ]

        # Crear archivo Excel en memoria
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Payment Import Template")

        # Estilo para encabezado
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'align': 'center', 'valign': 'vcenter'})
        
        # Escribir encabezados en la primera fila
        for col, label in enumerate(field_labels):
            worksheet.write(0, col, label, header_format)

        # Añadir una línea de ejemplo
        example_data = [
            1000.0,                   # Amount
            "2024-03-19 08:00:00",    # Date
            "Bank",                   # Journal
            "Proveedor ABC",           # Customer/Vendor
            "inbound",                # Payment Type (Can be 'inbound' or 'outbound')
            "PAY-12345"               # Number
        ]

        # Escribir los datos de ejemplo en la segunda fila
        for col, value in enumerate(example_data):
            worksheet.write(1, col, value)

        # Ajustar automáticamente el tamaño de las columnas para que todo el texto sea visible
        for col in range(len(field_labels)):
            column_width = max(len(field_labels[col]), max(len(str(example_data[col])) for example_data in [example_data]))
            worksheet.set_column(col, col, column_width)

        # Cerrar y preparar el archivo
        workbook.close()
        output.seek(0)

        # Crear adjunto en Odoo
        attachment = self.env['ir.attachment'].create({
            'name': 'payment_import_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Retornar archivo para descarga
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
